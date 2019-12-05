import os
from datetime import datetime

from django.http import HttpResponse
from openpyxl import load_workbook, Workbook

from kbd_aps_data_tool import settings

from django.shortcuts import render, redirect


def index(request):
    # return render(request, 'index.html')
    return redirect('converter')


def converter(request):
    context = {'files': {
        'product_attribute': [{'name': 'a.xlsx', 'value': 'data/a-uuid.a.xlsx'}],
        'item_project': [{'name': 'a.xlsx', 'value': 'data/a-uuid.a.xlsx'}]
    }}
    return render(request, 'converter.html', context=context)


def converter_upload(request):
    files = request.FILES.getlist("file")
    if files == None or len(files) <= 0:
        return HttpResponse('文件不能为空', status=400)
    for file in files:
        if file.content_type != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            return HttpResponse('请上传EXCEL文件,文件类型为.xlsx', status=400)
        else:
            wb = load_workbook(filename=file, read_only=True, data_only=True)
            path = os.path.join(settings.BASE_DIR, 'output')
            if not os.path.exists(path):
                os.makedirs(path, exist_ok=True)
            filepath = os.path.join(path, file)
            wb.save(filepath)


def convert(request):
    project_attribute = '/home/chen/Desktop/kbd_aps_data_tool/medias/机种属性.xlsx'
    item_project = '/home/chen/Desktop/kbd_aps_data_tool/medias/物料机种.xlsx'
    project_attribute_wb = load_workbook(project_attribute)
    item_project_wb = load_workbook(item_project)
    sheet1 = project_attribute_wb[project_attribute_wb.sheetnames[0]]
    sheet2 = item_project_wb[item_project_wb.sheetnames[0]]
    # 表总行数
    max_row = sheet2.max_row
    # 表总列数
    max_col = 4
    item_project_list = []
    for x in range(1, max_row + 1):
        list1 = []
        for y in range(1, max_col + 1):
            # 获取表中x行y列的值
            cell_data = sheet2.cell(row=x, column=y).value
            if cell_data:
                list1.append(cell_data)
        item_project_list.append(list1)
    item_project_dict = {}
    a = item_project_list[0].index('工序')
    b = item_project_list[0].index(('物料编码'))
    c = item_project_list[0].index(('机种/项目'))
    d = item_project_list[0].index(('地点编码'))
    item_project_list.pop(0)
    for k in item_project_list:
        if k:
            try:
                item_project_dict[(k[a], k[d], k[c])]

                item_project_dict[(k[a], k[d], k[c])].append(k[b])
            except:
                item_project_dict[(k[a], k[d], k[c])] = [k[b]]

    max_row1 = sheet1.max_row
    max_col1 = 7
    project_attribute_list = []
    for x in range(1, max_row1 + 1):
        list2 = []
        for y in range(1, max_col1 + 1):
            # 获取表中x行y列的值
            cell_data = sheet1.cell(row=x, column=y).value
            if cell_data:
                list2.append(cell_data)
        project_attribute_list.append(list2)
    project_attribute_dict = {}
    # ['机种/项目', '工序', '资源编码', 'UPH', '总人力', '单位人工工时 (S/PCS)', '地点']
    a = project_attribute_list[0].index(('机种/项目'))
    b = project_attribute_list[0].index(('工序'))
    c = project_attribute_list[0].index(('资源编码'))
    d = project_attribute_list[0].index(('UPH'))
    e = project_attribute_list[0].index(('总人力'))
    f = project_attribute_list[0].index(('单位人工工时 (S/PCS)'))
    g = project_attribute_list[0].index(('地点'))
    project_attribute_list.pop(0)
    for k in project_attribute_list:
        if k:
            try:
                project_attribute_dict[(k[b], k[g], k[a])]

                project_attribute_dict[(k[b], k[g], k[a])].append({k[c]: (k[d], k[f])})

            except:
                project_attribute_dict[(k[b], k[g], k[a])] = [{k[c]: (k[d], k[f])}]
    # operation_head =
    operation_list = [['代码', '地点编码', '物料编码']]
    operation_resource_list = [['工序编码', '资源编码', '地点编码', '物料编码', '标准UPH', '单位人工工时', '生产批量', '资源占用数量']]
    for k, v in item_project_dict.items():

        for i in v:
            list1 = [i + '_' + k[0], k[1], k[2]]
            operation_list.append(list1)
            try:
                m = project_attribute_dict[k]
                for z in m:
                    for e, f in z.items():
                        list2 = [list1[0], e, k[1], k[2], f[0], f[1], None, None]
                        operation_resource_list.append(list2)
            except:
                pass
    now = datetime.now()
    date_str = datetime.strftime(now, '%Y%m%d-%H%M%S')
    operation_name = '工序表' + date_str + '.xlsx'
    operation_resource_name = '工序资源表' + date_str + '.xlsx'
    operation_file_path = save_excel(operation_list, operation_name, '工序表')
    operation_resource_file_path = save_excel(operation_resource_list, operation_resource_name, '工序资源表')
    context = {'message': '处理成功',
               'files': [operation_file_path, operation_resource_file_path]}

    return render(request, 'convert_result.html', context=context)


def save_excel(list, filepath, title):
    wb = Workbook()
    # 默认表sheet1
    ws1 = wb.active
    # ws2 = wb.active
    # 更改表名
    ws1.title = title
    i = 1
    for k in list:
        j = 1
        for h in k:
            ws1.cell(row=i, column=j, value=h)
            j += 1
        i += 1
    # 保存于本地
    path = os.path.join(settings.BASE_DIR, 'input')
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)
    filepath = os.path.join(path, filepath)
    wb.save(filepath)
    return filepath
# def converter_upload(request):
#     files = request.FILES
#     return redirect('converter')
